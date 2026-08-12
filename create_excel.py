import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "接口测试数据"

# 写入表头
ws['A1'] = "接口类型"
ws['B1'] = "接口地址"
ws['C1'] = "期望状态码"

# 写入数据（模拟你的测试用例）
test_data = [
    ("GET", "/posts/1", 200),
    ("GET", "/posts/2", 200),
    ("POST", "/posts", 201),
    ("PUT", "/posts/1", 200),
]

for row_idx, (method, endpoint, code) in enumerate(test_data, start=2):
    ws.cell(row=row_idx, column=1, value=method)
    ws.cell(row=row_idx, column=2, value=endpoint)
    ws.cell(row=row_idx, column=3, value=code)

wb.save("data/test_data.xlsx")
print("✅ Excel 文件已生成：data/test_data.xlsx")